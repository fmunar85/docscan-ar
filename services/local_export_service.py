"""
Servicio de exportación local – DocScan AR
==========================================
Genera archivos XLSX y CSV descargables con formato profesional.

COMPROBANTE:
  • RESUMEN  → 1 fila por GRUPO con cantidad y costo sumados
  • DETALLE  → todas las líneas, columna Grupo con color alternado por grupo

Otros tipos: una hoja, cabecera coloreada, anchos auto, fila totales.
"""
import csv
import re
from datetime import datetime
from io import BytesIO, StringIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.sheets_service import HEADERS


# ─────────────────────────────────────────────────────────────────────────────
# Helpers generales
# ─────────────────────────────────────────────────────────────────────────────

def _now_ba() -> str:
    return datetime.now(ZoneInfo("America/Argentina/Buenos_Aires")).strftime(
        "%d/%m/%Y %H:%M"
    )


def _clean_filename(value: str) -> str:
    if not value:
        return "documento"
    cleaned = re.sub(r"[^\w\-\.]+", "_", str(value).strip(), flags=re.UNICODE)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned[:80] or "documento"


def _doc_number(data: dict, doc_type: str) -> str:
    dt = doc_type.upper()
    if dt == "FACTURA":
        return data.get("numero", "")
    if dt == "REMITO":
        return data.get("orden_salida", data.get("numero", ""))
    if dt == "COMPROBANTE":
        return data.get("comprobante_numero", "")
    return data.get("numero", "")


# ─────────────────────────────────────────────────────────────────────────────
# Estilos Excel
# ─────────────────────────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT  = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_INFO_FILL  = PatternFill("solid", fgColor="D6E4F0")
_INFO_FONT  = Font(bold=True, size=11, name="Calibri", color="1F4E79")
_INFO_ALIGN = Alignment(horizontal="left", vertical="center")

_TOT_FILL  = PatternFill("solid", fgColor="FFF2CC")
_TOT_FONT  = Font(bold=True, size=10, name="Calibri")

_GRP_FILLS = [
    PatternFill("solid", fgColor="EBF5FB"),   # azul muy claro
    PatternFill("solid", fgColor="EAFAF1"),   # verde muy claro
    PatternFill("solid", fgColor="FEF9E7"),   # amarillo muy claro
    PatternFill("solid", fgColor="FDEDEC"),   # rosa muy claro
    PatternFill("solid", fgColor="F4ECF7"),   # violeta muy claro
]
_GRP_FONT_BOLD = Font(bold=True, size=10, name="Calibri")
_GRP_HDR_FILLS = [
    PatternFill("solid", fgColor="2E86C1"),   # azul
    PatternFill("solid", fgColor="1E8449"),   # verde
    PatternFill("solid", fgColor="D4AC0D"),   # amarillo oscuro
    PatternFill("solid", fgColor="C0392B"),   # rojo
    PatternFill("solid", fgColor="7D3C98"),   # violeta
]

_NRM_FONT  = Font(size=9, name="Calibri")
_NRM_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)

_THIN   = Side(style="thin", color="BDD7EE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _auto_width(ws, min_w=8, max_w=55):
    for col_cells in ws.columns:
        length = min_w
        for cell in col_cells:
            try:
                length = max(length, min(len(str(cell.value or "")) + 2, max_w))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length


def _apply_cell(cell, fill=None, font=None, alignment=None, border=None):
    if fill:      cell.fill      = fill
    if font:      cell.font      = font
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


def _write_info_row(ws, row: int, ncols: int, label: str):
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1, value=label)
    for c in range(1, ncols + 1):
        _apply_cell(ws.cell(row=row, column=c),
                    fill=_INFO_FILL, font=_INFO_FONT, alignment=_INFO_ALIGN, border=_BORDER)
    ws.row_dimensions[row].height = 22


def _write_header_row(ws, row: int, headers: list):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        _apply_cell(cell, fill=_HDR_FILL, font=_HDR_FONT, alignment=_HDR_ALIGN, border=_BORDER)
    ws.row_dimensions[row].height = 28


def _write_total_row(ws, row: int, values: list):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        _apply_cell(cell, fill=_TOT_FILL, font=_TOT_FONT, alignment=_NRM_ALIGN, border=_BORDER)
    ws.row_dimensions[row].height = 18


def _write_plain_sheet(ws, headers, data_rows, info_label="", totals_row=None, freeze=True):
    ncols = len(headers)
    r = 1
    if info_label:
        _write_info_row(ws, r, ncols, info_label)
        r += 1
    _write_header_row(ws, r, headers)
    freeze_row = r + 1
    r += 1

    ALT = PatternFill("solid", fgColor="EAF2FB")
    NRM = PatternFill("solid", fgColor="FFFFFF")
    for i, row in enumerate(data_rows):
        fill = ALT if i % 2 else NRM
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            _apply_cell(cell, fill=fill, font=_NRM_FONT, alignment=_NRM_ALIGN, border=_BORDER)
        ws.row_dimensions[r].height = 15
        r += 1

    if totals_row:
        _write_total_row(ws, r, totals_row)

    if freeze:
        ws.freeze_panes = ws.cell(row=freeze_row, column=1).coordinate

    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Parsers de datos por tipo
# ─────────────────────────────────────────────────────────────────────────────

def _factura(data, now):
    h = HEADERS["FACTURAS"]
    row = [now, data.get("fecha",""), data.get("tipo_comprobante",""),
           data.get("numero",""), data.get("proveedor",""), data.get("cuit",""),
           data.get("subtotal",""), data.get("iva_porcentaje",""),
           data.get("iva_monto",""), data.get("total",""),
           data.get("cae",""), data.get("vencimiento_cae",""), data.get("observaciones","")]
    tot = [""] * len(h); tot[0] = "TOTAL"; tot[9] = data.get("total","")
    return h, [row], tot


def _remito(data, now):
    h = HEADERS["REMITOS"]
    raw_art = data.get("articulos","")
    lineas = [l.strip() for l in raw_art.splitlines() if l.strip()] or [""]
    base = [now, data.get("fecha",""), data.get("orden_salida",""),
            data.get("pack",""), data.get("origen_nombre",""),
            data.get("origen_direccion",""), data.get("origen_ciudad",""),
            data.get("origen_telefono",""), data.get("destinatario",""),
            data.get("destino_direccion",""), data.get("destino_ciudad",""),
            data.get("destino_telefono",""), data.get("peso_total",""), data.get("observaciones","")]
    rows = []
    for line in lineas:
        parts = [p.strip() for p in line.split("|")]
        while len(parts) < 6: parts.append("")
        rows.append(base + parts[:6])
    tot = [""] * len(h)
    tot[0] = f"TOTAL — {len(rows)} artículo(s)"
    tot[12] = data.get("peso_total","")
    return h, rows, tot


def _ticket(data, now):
    h = HEADERS["TICKETS"]
    row = [now, data.get("fecha",""), data.get("comercio",""),
           data.get("concepto",""), data.get("categoria",""),
           data.get("monto",""), data.get("observaciones","")]
    tot = [""] * len(h); tot[0] = "TOTAL"; tot[5] = data.get("monto","")
    return h, [row], tot


def _comprobante_parsed(data, now):
    """
    Parsea comp_resumen_lineas (Grupo|Cant|Costo) y comp_detalle_lineas
    (Grupo|Tipo|LP|Cant|Detalle|Serie|Costo|Raw).
    Devuelve estructuras listas para el XLSX.
    """
    res_h = HEADERS["COMP_RESUMEN"]
    det_h = HEADERS["COMP_DETALLE"]

    res_lineas = [l.strip() for l in data.get("comp_resumen_lineas","").splitlines() if l.strip()] or [""]
    det_lineas = [l.strip() for l in data.get("comp_detalle_lineas","").splitlines() if l.strip()] or [""]

    meta_base = [
        now,
        data.get("fecha",""),
        data.get("comprobante_numero",""),
        data.get("cliente",""),
        data.get("direccion",""),
        data.get("cantidad_total",""),
        data.get("total_usd",""),
    ]

    # ── RESUMEN: 1 fila por grupo ──
    res_rows = []
    for line in res_lineas:
        parts = [p.strip() for p in line.split("|")]
        while len(parts) < 3: parts.append("")
        grupo, cant, costo = parts[0], parts[1], parts[2]
        res_rows.append(meta_base + [grupo, cant, costo, data.get("observaciones","")])

    # Totales resumen
    try:
        total_cant = sum(int(r[9]) for r in res_rows if r[9].strip().isdigit())
    except Exception:
        total_cant = ""
    res_tot = [""] * len(res_h)
    res_tot[0] = f"TOTAL — {len(res_rows)} grupo(s)"
    res_tot[5] = data.get("cantidad_total","")
    res_tot[6] = data.get("total_usd","")
    res_tot[9] = str(total_cant) if total_cant else ""

    # ── DETALLE: todas las líneas + grupo ──
    det_base = [
        now,
        data.get("fecha",""),
        data.get("comprobante_numero",""),
        data.get("cliente",""),
        data.get("direccion",""),
    ]
    det_rows = []
    for line in det_lineas:
        parts = [p.strip() for p in line.split("|")]
        while len(parts) < 8: parts.append("")
        det_rows.append(det_base + parts[:8] + [data.get("observaciones","")])

    return res_h, res_rows, res_tot, det_h, det_rows


# ─────────────────────────────────────────────────────────────────────────────
# XLSX builder principal
# ─────────────────────────────────────────────────────────────────────────────

def _write_comprobante_resumen(ws, headers, res_rows, res_tot, info_label):
    """Hoja RESUMEN: 1 fila/grupo. Grupo column en negrita, totales amarillos."""
    ncols = len(headers)
    r = 1

    _write_info_row(ws, r, ncols, info_label)
    r += 1

    _write_header_row(ws, r, headers)
    ws.freeze_panes = ws.cell(row=r + 1, column=1).coordinate
    r += 1

    # Columna de Grupo = índice 7 (0-based), que es la col 8 (1-based) en la hoja
    # headers: FechReg, FechaDoc, CompNro, Cliente, Direcc, CantTotal, TotalUSD, Grupo, Cant.Grupo, CostoGrupo, Obs
    grupo_col_idx = next((i+1 for i, h in enumerate(headers) if "Grupo" in h), 8)

    GRP_COLORS = [
        ("2471A3", "D6EAF8"),  # azul
        ("1E8449", "D5F5E3"),  # verde
        ("B7950B", "FEF9E7"),  # dorado
        ("A04000", "FDEBD0"),  # naranja
        ("6C3483", "F4ECF7"),  # violeta
        ("117A65", "D1F2EB"),  # teal
        ("2E4053", "EBF5FB"),  # gris azul
    ]

    prev_group = None
    color_idx  = -1

    for row_data in res_rows:
        grupo = row_data[grupo_col_idx - 1] if len(row_data) >= grupo_col_idx else ""
        if grupo != prev_group:
            color_idx  = (color_idx + 1) % len(GRP_COLORS)
            prev_group = grupo
        hdr_hex, body_hex = GRP_COLORS[color_idx]
        body_fill = PatternFill("solid", fgColor=body_hex)
        hdr_fill  = PatternFill("solid", fgColor=hdr_hex)

        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            is_grupo_col = (c == grupo_col_idx)
            cell.fill      = hdr_fill if is_grupo_col else body_fill
            cell.font      = Font(bold=True, size=10, color="FFFFFF", name="Calibri") if is_grupo_col else _NRM_FONT
            cell.alignment = _NRM_ALIGN
            cell.border    = _BORDER
        ws.row_dimensions[r].height = 16
        r += 1

    _write_total_row(ws, r, res_tot)
    _auto_width(ws)


def _write_comprobante_detalle(ws, headers, det_rows, info_label):
    """Hoja DETALLE: todas las líneas, color alternado por grupo."""
    ncols = len(headers)
    r = 1

    _write_info_row(ws, r, ncols, info_label)
    r += 1

    _write_header_row(ws, r, headers)
    ws.freeze_panes = ws.cell(row=r + 1, column=1).coordinate
    r += 1

    # Grupo col = index 5 in headers (0-based), col 6 (1-based)
    grupo_col_idx = next((i+1 for i, h in enumerate(headers) if h == "Grupo"), 6)

    GRP_PAIRS = [
        ("D6EAF8", "EBF5FB"),
        ("D5F5E3", "EAFAF1"),
        ("FEF9E7", "FFFDE7"),
        ("FDEBD0", "FEF5EC"),
        ("F4ECF7", "F9F0FB"),
        ("D1F2EB", "E8F8F5"),
    ]
    prev_group = None
    color_idx  = -1

    for row_data in det_rows:
        grupo = row_data[grupo_col_idx - 1] if len(row_data) >= grupo_col_idx else ""
        if grupo != prev_group:
            color_idx  = (color_idx + 1) % len(GRP_PAIRS)
            prev_group = grupo
        grp_hex, row_hex = GRP_PAIRS[color_idx]
        grp_fill = PatternFill("solid", fgColor=grp_hex)
        row_fill = PatternFill("solid", fgColor=row_hex)

        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            is_grupo = (c == grupo_col_idx)
            cell.fill      = grp_fill if is_grupo else row_fill
            cell.font      = Font(bold=True, size=9, name="Calibri") if is_grupo else _NRM_FONT
            cell.alignment = _NRM_ALIGN
            cell.border    = _BORDER
        ws.row_dimensions[r].height = 14
        r += 1

    _auto_width(ws)


def _xlsx(data: dict, doc_type: str) -> bytes:
    dt  = doc_type.upper()
    now = _now_ba()
    wb  = Workbook()

    if dt == "COMPROBANTE":
        res_h, res_rows, res_tot, det_h, det_rows = _comprobante_parsed(data, now)
        comp    = data.get("comprobante_numero","")
        cliente = data.get("cliente","")
        fecha   = data.get("fecha","")
        info    = f"Comprobante Nº {comp}  |  Cliente: {cliente}  |  Fecha: {fecha}"

        ws1 = wb.active
        ws1.title = "Comp. Resumen"
        _write_comprobante_resumen(ws1, res_h, res_rows, res_tot, info)

        ws2 = wb.create_sheet(title="Comp. Detalle")
        _write_comprobante_detalle(ws2, det_h, det_rows, info)

    elif dt == "FACTURA":
        h, rows, tot = _factura(data, now)
        info = f"Factura Nº {data.get('numero','')}  |  {data.get('proveedor','')}  |  Fecha: {data.get('fecha','')}"
        ws = wb.active; ws.title = "Factura"
        _write_plain_sheet(ws, h, rows, info_label=info, totals_row=tot)

    elif dt == "REMITO":
        h, rows, tot = _remito(data, now)
        info = f"Remito Nº {data.get('orden_salida','')}  |  {data.get('destinatario','')}  |  Fecha: {data.get('fecha','')}"
        ws = wb.active; ws.title = "Remito"
        _write_plain_sheet(ws, h, rows, info_label=info, totals_row=tot)

    else:  # TICKET
        h, rows, tot = _ticket(data, now)
        info = f"Ticket  |  {data.get('comercio','')}  |  Fecha: {data.get('fecha','')}"
        ws = wb.active; ws.title = "Ticket"
        _write_plain_sheet(ws, h, rows, info_label=info, totals_row=tot)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# CSV builder
# ─────────────────────────────────────────────────────────────────────────────

def _csv(data: dict, doc_type: str) -> bytes:
    dt  = doc_type.upper()
    now = _now_ba()
    out = StringIO()
    wr  = csv.writer(out, quoting=csv.QUOTE_MINIMAL)

    if dt == "COMPROBANTE":
        res_h, res_rows, res_tot, det_h, det_rows = _comprobante_parsed(data, now)
        comp = data.get("comprobante_numero","")

        wr.writerow([f"=== COMP. RESUMEN — Comprobante Nº {comp} ==="])
        wr.writerow(res_h)
        wr.writerows(res_rows)
        wr.writerow(res_tot)
        wr.writerow([])
        wr.writerow([f"=== COMP. DETALLE — Comprobante Nº {comp} ==="])
        wr.writerow(det_h)
        wr.writerows(det_rows)

    elif dt == "FACTURA":
        h, rows, tot = _factura(data, now); wr.writerow(h); wr.writerows(rows); wr.writerow(tot)
    elif dt == "REMITO":
        h, rows, tot = _remito(data, now);  wr.writerow(h); wr.writerows(rows); wr.writerow(tot)
    else:
        h, rows, tot = _ticket(data, now);  wr.writerow(h); wr.writerows(rows); wr.writerow(tot)

    return out.getvalue().encode("utf-8-sig")


# ─────────────────────────────────────────────────────────────────────────────
# Punto de entrada público
# ─────────────────────────────────────────────────────────────────────────────

def build_local_export(data: dict, doc_type: str, file_format: str) -> dict:
    dt          = (doc_type or "FACTURA").upper()
    file_format = (file_format or "xlsx").lower()
    data        = data or {}

    number    = _clean_filename(_doc_number(data, dt))
    base_name = f"{dt}_{number}" if number and number != "documento" else dt

    if file_format == "csv":
        return {
            "success":  True,
            "bytes":    _csv(data, dt),
            "mimetype": "text/csv",
            "filename": f"{base_name}.csv",
        }

    return {
        "success":  True,
        "bytes":    _xlsx(data, dt),
        "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "filename": f"{base_name}.xlsx",
    }
